import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from typing import List, Dict, Any, Union
import os
import tempfile
import unittest

class ExcelUtil:
    @staticmethod
    def read_excel(file_path: str, sheet_name: str = None) -> List[Dict[str, Any]]:
        """
        读取Excel文件并返回数据列表
        
        :param file_path: Excel文件路径
        :param sheet_name: 工作表名称,如果为None则读取第一个工作表
        :return: 包含每行数据的字典列表
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            if sheet_name:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]
            data = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = {headers[i]: value for i, value in enumerate(row)}
                data.append(row_data)

            return data
        except FileNotFoundError:
            raise ValueError(f"文件 {file_path} 不存在。")
        except InvalidFileException:
            raise ValueError(f"无法打开文件 {file_path}。请确保它是一个有效的Excel文件。")
        except KeyError:
            raise ValueError(f"工作表 '{sheet_name}' 不存在。")
        except Exception as e:
            raise Exception(f"读取Excel文件时发生错误: {str(e)}")

    @staticmethod
    def write_excel(data: List[Dict[str, Any]], file_path: str, sheet_name: str = "Sheet1"):
        """
        将数据写入Excel文件
        
        :param data: 要写入的数据,每个字典代表一行
        :param file_path: 要创建或覆盖的Excel文件路径
        :param sheet_name: 工作表名称
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        # 写入表头
        headers = list(data[0].keys())
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        # 写入数据
        for row, row_data in enumerate(data, start=2):
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=row, column=col, value=row_data.get(header))

        try:
            workbook.save(file_path)
        except PermissionError:
            raise PermissionError(f"无法写入文件 {file_path}。文件可能已被打开或您没有写入权限。")
        except Exception as e:
            raise Exception(f"写入Excel文件时发生错误: {str(e)}")

    @staticmethod
    def append_to_excel(data: List[Dict[str, Any]], file_path: str, sheet_name: str = None):
        """
        向现有Excel文件追加数据
        
        :param data: 要追加的数据,每个字典代表一行
        :param file_path: 现有Excel文件的路径
        :param sheet_name: 工作表名称,如果为None则使用第一个工作表
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件 {file_path} 不存在。")

            workbook = openpyxl.load_workbook(file_path)
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    sheet = workbook.create_sheet(sheet_name)
                else:
                    sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            # 如果工作表为空,写入表头
            if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
                headers = list(data[0].keys())
                for col, header in enumerate(headers, start=1):
                    sheet.cell(row=1, column=col, value=header)
                start_row = 2
            else:
                headers = [cell.value for cell in sheet[1]]
                start_row = sheet.max_row + 1

            # 追加数据
            for row, row_data in enumerate(data, start=start_row):
                for col, header in enumerate(headers, start=1):
                    sheet.cell(row=row, column=col, value=row_data.get(header))

            workbook.save(file_path)
        except InvalidFileException:
            raise ValueError(f"无法打开文件 {file_path}。请确保它是一个有效的Excel文件。")
        except PermissionError:
            raise PermissionError(f"无法写入文件 {file_path}。文件可能已被打开或您没有写入权限。")
        except Exception as e:
            raise Exception(f"追加数据到Excel文件时发生错误: {str(e)}")

if __name__ == "__main__":
    class TestExcelUtil(unittest.TestCase):
        def setUp(self):
            self.test_data = [
                {"Name": "Alice", "Age": 30, "City": "New York"},
                {"Name": "Bob", "Age": 25, "City": "London"},
                {"Name": "Charlie", "Age": 35, "City": "Paris"}
            ]
            self.temp_dir = tempfile.mkdtemp()
            self.test_file = os.path.join(self.temp_dir, "test.xlsx")

        def tearDown(self):
            if os.path.exists(self.test_file):
                os.remove(self.test_file)
            os.rmdir(self.temp_dir)

        def test_write_and_read_excel(self):
            print("测试写入和读取Excel文件...")
            ExcelUtil.write_excel(self.test_data, self.test_file)
            self.assertTrue(os.path.exists(self.test_file))
            print("文件写入成功:", self.test_file)

            read_data = ExcelUtil.read_excel(self.test_file)
            self.assertEqual(len(read_data), len(self.test_data))
            for original, read in zip(self.test_data, read_data):
                self.assertDictEqual(original, read)
            print("文件读取成功，数据匹配")

        def test_append_to_excel(self):
            print("测试向Excel文件追加数据...")
            ExcelUtil.write_excel(self.test_data, self.test_file)
            append_data = [
                {"Name": "David", "Age": 28, "City": "Tokyo"},
                {"Name": "Eva", "Age": 33, "City": "Berlin"}
            ]
            ExcelUtil.append_to_excel(append_data, self.test_file)

            read_data = ExcelUtil.read_excel(self.test_file)
            self.assertEqual(len(read_data), len(self.test_data) + len(append_data))
            for original, read in zip(self.test_data + append_data, read_data):
                self.assertDictEqual(original, read)
            print("数据追加成功，数据匹配")

        def test_read_nonexistent_file(self):
            print("测试读取不存在的文件...")
            with self.assertRaises(ValueError):
                ExcelUtil.read_excel("nonexistent.xlsx")
            print("正确处理不存在的文件")

        def test_read_invalid_sheet(self):
            print("测试读取不存在的工作表...")
            ExcelUtil.write_excel(self.test_data, self.test_file)
            with self.assertRaises(ValueError):
                ExcelUtil.read_excel(self.test_file, "InvalidSheet")
            print("正确处理不存在的工作表")

    unittest.main()